# to work with excel, need to install openpyxl in command prompt and add the module library in PyCharm
# AGENDA :
# 1. Read data from Excel
# 2. Writing data into Excel
# 3. Data Driven Package1 case

import openpyxl
'''
# 1. READ DATA from EXCEL
# need to specify location of excel file
path = "C:/Users/Diyanah/Documents/test automation.xlsx"

# need to open the excel file and store it in workbook object
workbook = openpyxl.load_workbook(path)

# extract active sheets in the docs and store it in sheet object
sheet = workbook.active
# sheet1 = workbook.get_sheet_by_name("Sheet1")   # extract specific sheet, say sheet 1

# now we need to read the data using for loop
# need to find how many row and column first
rows = sheet.max_row    # 13
cols = sheet.max_column # 4
print("Rows :", rows)
print("Columns :", cols)

for r in range(1, rows + 1):
    for c in range(1, cols + 1):
        val = sheet.cell(row=r, column=c).value    # print the value of the sheet of the specific r,c
        print(val, end="    ")
    print()


# 2. WRITE DATA into EXCEL
path = "C:/Users/Diyanah/Documents/test 2.xlsx"
workbook = openpyxl.load_workbook(path)
sheet = workbook.active

# say we want to write data of 3 col and 5 row
for r in range(1, 6):
    for c in range(1, 4):
        sheet.cell(row=r, column=c).value = str(r) + str(c)

# save the workbook
workbook.save(path)


# 3. DATA DRIVEN TESTING in EXCEL
# - open the file
# - find max row and column
# - read the data
# - write data
# to make it simpler, lets define a module with all the excel functions in it, lets call them XLUtils

# say we wanna test the login of a website for a set of data
from XLUtils import *
from selenium import webdriver
from time import sleep

driver = webdriver.Chrome(executable_path="C:/Users/Diyanah/Downloads/chromedriver_win32/chromedriver.exe")

driver.get("https://app.kakitangan.com/login")

# now we need to read the data
# first, need to know how many rows and column
path = "C:/Users/Diyanah/Documents/test 1.xlsx"

rows = getRowCount(path, "Sheet1")
cols = getColumnCount(path, "Sheet1")

# now we read the coloumn 1 that contains user name, and column 2 for password
for r in range(2, rows + 1):
    username = readData(path, "Sheet1", r, 1)       # column 1 for row 2-end
    password = readData(path, "Sheet1", r, 2)       # column 2 for row 2-end

    driver.find_element_by_name("username").send_keys(username)     # login with the details
    driver.find_element_by_name("password").send_keys(password)
    driver.find_element_by_css_selector("button.btn.btn-primary.form-control").click()

    sleep(10)

    if driver.title == "Leave Application | Kakitangan":        # check if able to login
        writeData(path, "Sheet1", r, 3, "Success")
        driver.back()  # go back to login page for next iteration
        sleep(10)
    else:
        writeData(path, "Sheet1", r, 3, "Fail")

    driver.find_element_by_name("username").clear()
    driver.find_element_by_name("password").clear()
'''

