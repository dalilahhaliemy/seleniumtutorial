import openpyxl


def getRowCount(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetName)
    return sheet.max_row


def getColumnCount(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetName)
    return sheet.max_column


def readData(file, sheetName, rowNum, colNum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetName)
    return sheet.cell(row=rowNum, column=colNum).value


def writeData(file, sheetName, rowNum, colNum, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetName)
    sheet.cell(row=rowNum, column=colNum).value = data
    workbook.save(file)



class Excel:

    def load(self, path):
        openpyxl.load_workbook(path)

        # extract active sheets in the docs and store it in sheet object
        sheet = workbook.active
        # sheet1 = workbook.get_sheet_by_name("Sheet1")   # extract specific sheet, say sheet 1

        # now we need to read the data using for loop
        # need to find how many row and column first
        rows = sheet.max_row  # 13
        cols = sheet.max_column  # 4
        print(rows)
        print(cols)

        for r in range(1, rows + 1):
            for c in range(1, cols + 1):
                val = sheet.cell(row=r, column=c).value  # print the value of the sheet of the specific r,c
                print(val, end="    ")
            print()

        # 2. WRITE DATA into EXCEL
        path = "C:/Users/Diyanah/Documents/test 2.xlsx"
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.active

        # say we want to write data of 3 col and 5 row
        for r in range(1, 6):
            for c in range(1, 4):
                sheet.cell(row=r, column=c).value = str(r) + str(c)

        # save the workbook
        workbook.save(path)
